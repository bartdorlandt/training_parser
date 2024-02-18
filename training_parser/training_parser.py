"""Training parser main script."""
from pathlib import Path

import openpyxl

PATHS = list[Path]
parse_training_years = [2023]

mapper = {
    "lopen": [],
    "klimmen": [],
    "eind spel": [],
}


def generate_year_paths(base_path: Path, years: list[int]) -> PATHS:
    """Find the training years in the base path."""
    return [base_path.joinpath(f"trainingen {year}") for year in years]


def get_files(paths: PATHS) -> PATHS:
    """Get the xls files from the paths."""
    new_paths = []
    for path in paths:
        new_paths.extend(path.glob("*.xlsx"))
    return new_paths


def parse_xlsx_2024_file(path: Path) -> None:
    """Parse the xlsx file, 2024 format."""
    keyword = ""
    interesting_column = ""
    m = None
    triggers = ("LOPEN", "KLIMMEN", "eind spel")
    workbook = openpyxl.load_workbook(path)
    if "Basis training" not in workbook.sheetnames:
        print(f"Sheet 'Basis training' not found in {path}")
        return
    basis_sheet = workbook.get_sheet_by_name("Basis training")
    for row in basis_sheet.rows:
        for cell in row:
            if cell.value in triggers:
                keyword = cell.value.lower()
                print(f"{path}; {keyword}; {cell}")
                m = mapper[keyword]
                interesting_column = cell.column
            if (
                keyword
                and cell.column == interesting_column
                and cell.value
                and cell.value not in triggers
            ):
                m.append(cell.value)


def parse_xlsx_2023_file(path: Path) -> list[str]:  # sourcery skip: de-morgan
    """Parse the xlsx file, 2023 format."""
    interesting_column = "C"
    workbook = openpyxl.load_workbook(path)
    sheet_name = "Basis training"
    if sheet_name not in workbook.sheetnames:
        # print(f"Sheet '{sheet_name}' not found in {path}, trying 'Blad1'.")
        if "Blad1" in workbook.sheetnames:
            sheet_name = "Blad1"
        else:
            print(f"Sheet 'Blad1' not found in {path}")
            return []
    basis_sheet = workbook.get_sheet_by_name(sheet_name)
    started = False
    grouping = {}

    for row in basis_sheet.rows:
        for cell in row:
            if not (cell.column_letter == "A" and cell.value == "1#") and not started:
                continue
            started = True
            if cell.column_letter == "A" and cell.value and "#" in cell.value:
                number = cell.value.split("#")[0]
                grouping[number] = grouping.get(number, [])
            if cell.column_letter == interesting_column and cell.value:
                grouping[number].append(cell.value)

    return [" ".join(v) for v in grouping.values()]
